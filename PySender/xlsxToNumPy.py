import openpyxl
import sys
import numpy as np

if len(sys.argv) != 3:
    print("USAGE: python3 xlsxToNumPy.py <filepath> <nationalPrefix>")
    sys.exit()
elif not len(sys.argv[2]) == 2:
    print("ERROR: nationalPrefix should be 2 digits")
    sys.exit()
else:
    path = sys.argv[1]
    nationalPrefix = sys.argv[2]

wb = openpyxl.load_workbook(path)
sheet = wb[wb.sheetnames[0]]  # always select the first sheet

columnA = sheet["A"]  # always select the first column
count = 0
numbers = np.empty(sheet.max_row, dtype='object')
for cell in columnA:
    value = str(cell.value)
    if len(value) == 10 and not np.any(numbers[:] == (nationalPrefix + value)):
        numbers[count] = nationalPrefix + value
        count += 1
    elif len(value) == 12 and value.startswith(nationalPrefix) and not np.any(numbers[:] == value):
        numbers[count] = value
        count += 1
numbersReduced = numbers[0:count]
np.save("./numbers", numbersReduced)
print("Found ", len(numbersReduced), " numbers in ",
      path, " and saved to numbers.npy")
